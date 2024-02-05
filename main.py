import json
import openpyxl

def replace_occup_by_2_col(json_file_path, excel_file_path):
    # Чтение данных из JSON файла
    with open(json_file_path, 'r') as json_file:
        data = json.load(json_file)

    # Чтение данных из Excel файла
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Создание словаря для сопоставления значений из первого и второго столбца Excel
    replace_dict = dict((sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=2).value) for row in range(1, sheet.max_row + 1))

    # Замена значений поля "occupations" в JSON файле
    for item in data:
        if item.get('occupations') in replace_dict:
            item['occupations'] = replace_dict[item['occupations']]

    # Сохранение изменений в JSON файле
    with open(json_file_path, 'w') as json_file:
        json.dump(data, json_file, indent=4)

# Пример использования
replace_occup_by_2_col('occupations.json', 'to_replace.xlsx')